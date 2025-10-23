# common/xlsx_fill.py
from __future__ import annotations

import io
import os
import re
import unicodedata
from datetime import date, datetime
from typing import Optional, Dict, List, Tuple
from pathlib import Path

from dateutil.relativedelta import relativedelta
import pandas as pd
import openpyxl
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.drawing.image import Image as XLImage
try:
    # Pour ancrer précisément l'image sur une plage (TwoCellAnchor)
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
except Exception:
    AnchorMarker = TwoCellAnchor = None  # fallback si version openpyxl ancienne

from reportlab.lib.utils import ImageReader  # utilisé par la partie PDF
def _has_pillow() -> bool:
    try:
        import PIL  # noqa: F401
        return True
    except Exception:
        return False
def _add_logo(ws, path: Path | None, anchor_cell: str, max_w: int, max_h: int):
    """Ajoute un logo ancré sans déformer l'image (no-op si chemin invalide)."""
    try:
        if not path or not path.exists():
            print(f"[xlsx_fill] Logo introuvable pour ancre {anchor_cell} -> {path}")
            return
        from PIL import Image as PILImage
        from openpyxl.drawing.image import Image as XLImage

        with PILImage.open(path) as im:
            ow, oh = im.size

        scale = min(max_w / ow, max_h / oh, 1.0)  # pas d'upscale
        img = XLImage(str(path))
        img.width  = max(1, int(round(ow * scale)))
        img.height = max(1, int(round(oh * scale)))
        ws.add_image(img, anchor_cell)
        print(f"[xlsx_fill] Logo ajouté: {path.name} -> {anchor_cell} ({img.width}x{img.height}px)")
    except Exception as e:
        print(f"[xlsx_fill] ERREUR ajout logo {path} @ {anchor_cell}: {e}")


# ======================================================================
#                        Normalisation & mapping goûts
# ======================================================================

def _norm_key(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("’", "'")
    s = re.sub(r"[\s\-_/]+", " ", s)
    return " ".join(s.split())

# Canonical -> libellé EXACT attendu par Excel (étends la liste si besoin)
EXCEL_LABEL_MAP = {
    _norm_key("Original"):               "K. Original",
    _norm_key("Menthe citron vert"):     "K. Menthe - Citron Vert",
    _norm_key("Gingembre"):              "K. Gingembre",
    _norm_key("Pamplemousse"):           "K. Pamplemousse",
    _norm_key("Mangue Passion"):         "K. Mangue - Passion",
    _norm_key("Menthe Poivree"):         "EP. Menthe Poivrée",
    _norm_key("Mélisse"):                "EP. Mélisse",
    _norm_key("Anis étoilée"):           "EP. Anis étoilée",
    _norm_key("Zeste d'agrumes"):        "EP. Zest d'agrumes",
    _norm_key("Pêche"):                  "IG. Pêche",
    _norm_key("Autre"):                  "Autre :",
}

def _to_excel_label(gout: str) -> str:
    return EXCEL_LABEL_MAP.get(_norm_key(gout), str(gout or ""))

# ======================================================================
#                        Utilitaires de chemin/asset
# ======================================================================

def _project_root() -> Path:
    """Racine du projet (= dossier parent de 'common')."""
    try:
        return Path(__file__).resolve().parents[1]
    except Exception:
        return Path(os.getcwd())

def _load_asset_bytes(rel_path: str) -> bytes | None:
    """
    Charge un fichier d'assets en bytes, peu importe le cwd.
    Essaie <racine>/<rel_path> puis <rel_path> si déjà absolu.
    """
    root = _project_root()
    candidates = [root / rel_path, Path(rel_path)]
    for p in candidates:
        try:
            if p.exists() and p.is_file():
                return p.read_bytes()
        except Exception:
            pass
    return None

# ======================================================================
#                         Utilitaires généraux
# ======================================================================

VOL_TOL = 0.02

def _is_close(a: float, b: float, tol: float = VOL_TOL) -> bool:
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False

# ----------- parse format depuis la colonne "Stock" (df_min) -----------
def _parse_format_from_stock(stock: str):
    s = str(stock or "")
    m_nb = re.search(r'(Carton|Pack)\s+de\s+(\d+)\s+Bouteilles?', s, flags=re.I)
    nb = int(m_nb.group(2)) if m_nb else None
    m_l = re.search(r'(\d+(?:[.,]\d+)?)\s*[lL]\b', s)
    vol = float(m_l.group(1).replace(",", ".")) if m_l else None
    if vol is None:
        m_cl = re.search(r'(\d+(?:[.,]\d+)?)\s*c[lL]\b', s)
        vol = float(m_cl.group(1).replace(",", "."))/100.0 if m_cl else None
    return nb, vol

# ----------- Agrégat STRICT depuis df_min (tableau affiché) -----------
def _agg_from_dfmin(df_min: pd.DataFrame, gout: str) -> Dict[str, Dict[str, int]]:
    out = {
        "33_fr":  {"cartons": 0, "bouteilles": 0},
        "33_niko":{"cartons": 0, "bouteilles": 0},
        "75x6":   {"cartons": 0, "bouteilles": 0},
        "75x4":   {"cartons": 0, "bouteilles": 0},
    }
    if df_min is None or not isinstance(df_min, pd.DataFrame) or df_min.empty:
        return out
    req = {"Produit","Stock","GoutCanon","Cartons à produire (arrondi)","Bouteilles à produire (arrondi)"}
    if any(c not in df_min.columns for c in req):
        return out

    df = df_min.copy()
    df = df[df["GoutCanon"].astype(str).str.strip() == str(gout).strip()]
    if df.empty:
        return out

    for _, r in df.iterrows():
        nb, vol = _parse_format_from_stock(r["Stock"])
        if nb is None or vol is None:
            continue
        ct = int(pd.to_numeric(r["Cartons à produire (arrondi)"], errors="coerce") or 0)
        bt = int(pd.to_numeric(r["Bouteilles à produire (arrondi)"], errors="coerce") or 0)
        prod_up = str(r["Produit"]).upper()

        if nb == 12 and _is_close(vol, 0.33):
            key = "33_niko" if "NIKO" in prod_up else "33_fr"
        elif nb == 6 and _is_close(vol, 0.75):
            key = "75x6"
        elif nb == 4 and _is_close(vol, 0.75):
            key = "75x4"
        else:
            continue

        out[key]["cartons"]    += ct
        out[key]["bouteilles"] += bt

    return out

# ======================================================================
#                   Outils sûrs d’écriture Excel (fusions)
# ======================================================================

def _safe_set_cell(ws, row: int, col: int, value, number_format: str | None = None):
    """
    Écrit une valeur *même si* (row,col) tombe dans une cellule fusionnée.
    - Si (row,col) est l'ancre: écrit directement.
    - Si c'est à l'intérieur d'une fusion (pas l'ancre): on unmerge -> écrit à l'ancre -> remerge.
    Neutralise 'MergedCell ... value is read-only'.
    """
    hit = None
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            hit = rng
            break

    if hit is None:
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if number_format:
            cell.number_format = number_format
        return

    a_row, a_col = hit.min_row, hit.min_col
    coord = hit.coord  # ex: "C12:H14"
    # si on est déjà sur l'ancre, pas besoin de dé-fusionner
    if row == a_row and col == a_col:
        cell = ws.cell(row=a_row, column=a_col)
        cell.value = value
        if number_format:
            cell.number_format = number_format
        return

    # sinon on force
    ws.unmerge_cells(coord)
    cell = ws.cell(row=a_row, column=a_col)
    cell.value = value
    if number_format:
        cell.number_format = number_format
    ws.merge_cells(coord)

# Ecrit via adresse A1 ("D10" …) en gérant les fusions
def _set(ws, addr: str, value, number_format: str | None = None):
    row, col = coordinate_to_tuple(addr)
    _safe_set_cell(ws, row, col, value, number_format)
    return f"{get_column_letter(col)}{row}"

def _addr(col: int, row: int) -> str:
    return f"{get_column_letter(col)}{row}"

# ======================================================================
#                         Insertion d'image ancrée
# ======================================================================

def _add_image_in_range(ws, img_path: Path, tl_addr: str, br_addr: str):
    """
    Insère une image et l'ancre sur la plage [tl_addr:br_addr] (ex: 'P29'->'X51').
    - Essaie TwoCellAnchor (précis).
    - Sinon, fallback ws.add_image(img, tl_addr) + redimension approx.
    Loggue ce qu'il fait pour aider au debug.
    """
    try:
        if not img_path or not img_path.exists():
            print(f"[xlsx_fill] Image introuvable: {img_path}")
            return

        # charge l'image (nécessite Pillow)
        img = XLImage(str(img_path))
        print(f"[xlsx_fill] Image OK: {img_path.name}")

        # ---------- 1) Tentative TwoCellAnchor (précise) ----------
        if AnchorMarker and TwoCellAnchor:
            try:
                tl_row, tl_col = coordinate_to_tuple(tl_addr)  # 1-based
                br_row, br_col = coordinate_to_tuple(br_addr)
                frm = AnchorMarker(col=tl_col - 1, colOff=0, row=tl_row - 1, rowOff=0)
                to  = AnchorMarker(col=br_col - 1, colOff=0, row=br_row - 1, rowOff=0)
                img.anchor = TwoCellAnchor(_from=frm, _to=to, editAs='oneCell')
                ws.add_image(img)
                print("[xlsx_fill] Image ancrée via TwoCellAnchor.")
                return
            except Exception as e:
                print(f"[xlsx_fill] TwoCellAnchor indisponible/échec: {e}")

        # ---------- 2) Fallback: ancre en coin supérieur gauche ----------
        ws.add_image(img, tl_addr)
        print(f"[xlsx_fill] Image ajoutée en {tl_addr} (fallback). Redimension approx...")

        # Redimension approx pour couvrir la plage (si possible)
        # Conversion approximative des largeurs/hauteurs Excel -> pixels :
        # - largeur colonne (unités Excel) ~ 7 pixels
        # - hauteur ligne (points) -> pixels ~ points * 96/72
        def _col_pixels(col_idx_1b: int) -> int:
            col_letter = get_column_letter(col_idx_1b)
            w = ws.column_dimensions[col_letter].width
            if w is None:
                w = 8.43  # défaut Excel
            return int(round(w * 7.0))

        def _row_pixels(row_idx_1b: int) -> int:
            h = ws.row_dimensions[row_idx_1b].height
            if h is None:
                h = 15  # points, défaut Excel
            return int(round(h * (96.0 / 72.0)))

        tl_r, tl_c = coordinate_to_tuple(tl_addr)
        br_r, br_c = coordinate_to_tuple(br_addr)
        width_px  = sum(_col_pixels(c) for c in range(tl_c, br_c + 1))
        height_px = sum(_row_pixels(r) for r in range(tl_r, br_r + 1))
        if width_px > 0 and height_px > 0:
            img.width, img.height = width_px, height_px
            print(f"[xlsx_fill] Redimension: {width_px}x{height_px}px")
    except Exception as e:
        print(f"[xlsx_fill] ERREUR insertion image: {e}")

# ======================================================================
#                    Fiche de production (Grande/Petite)
# ======================================================================

def fill_fiche_7000L_xlsx(
    template_path: str,
    semaine_du: date,
    ddm: date,
    gout1: str,
    gout2: Optional[str],
    df_calc,
    sheet_name: str | None = None,
    df_min=None,
) -> bytes:
    """
    Remplit UNIQUEMENT :
      - H8 = libellé Excel du goût (via _to_excel_label)
      - A20 = date de début de fermentation (semaine_du)
      - cellule à droite du libellé "DDM" = ddm
      - K15/M15/O15/Q15/S15 = quantités de cartons (33cL France, 33cL NIKO, 75cL x6, 75cL x4, 75cL NIKO x6)
      - Image schéma cuves en P29:X51 (orange pour Grande, bleu pour Petite)
    Sans toucher aux autres formules.
    """
    import openpyxl
    import pandas as pd

    # --- ouverture & sélection de la feuille ---
    wb = openpyxl.load_workbook(template_path, data_only=False, keep_vba=False)
    targets = [sheet_name] if sheet_name else ["Fiche de production 7000 L", "Fiche de production 7000L"]
    ws = None
    for nm in targets:
        if nm and nm in wb.sheetnames:
            ws = wb[nm]
            break
    if ws is None:
        ws = wb.active  # fallback
        
    # --- Mise en page : exactement 1 page en largeur, hauteur libre ---
    try:
        # Format et marges
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize   = ws.PAPERSIZE_A4
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top  = 0.5
        ws.page_margins.bottom = 0.5
    
        # Ajuster à 1 page de large (et autant de pages que nécessaire en hauteur)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0   # 0 = pas de contrainte de hauteur
    
        # Neutraliser tout ancien zoom forcé
        ws.page_setup.scale = 100
    
        # Confort visuel à l'impression
        ws.page_setup.horizontalCentered = True
        # (laisse verticalCentered=False pour ne pas perdre de place en haut)
    except Exception:
        pass

     # --- Schéma cuves : ancrage fixe + mise à l'échelle proportionnelle (pas d'étirement) ---
    try:
        from PIL import Image as PILImage  # Pillow requis
        from openpyxl.drawing.image import Image as XLImage

        root = _project_root()
        base = Path(template_path).stem.lower()

        # Fichier selon le modèle
        if "grande" in base:
            img_file = root / "assets" / "schema_cuve_orange.png"
        elif "petite" in base:
            img_file = root / "assets" / "schema_cuve_bleu.png"
        else:
            img_file = root / "assets" / "schema_cuve_orange.png"

        if img_file.exists():
            # 1) lis la taille d'origine (px)
            with PILImage.open(img_file) as im:
                orig_w, orig_h = im.size

            # 2) cadre max (ajuste si besoin)
            # -> sur ta capture, une image ~300-340 px de large passe bien
            MAX_W, MAX_H = 340, 320   # px

            # 3) scale factor pour conserver le ratio
            scale = min(MAX_W / orig_w, MAX_H / orig_h, 1.0)
            out_w = int(round(orig_w * scale))
            out_h = int(round(orig_h * scale))

            # 4) on ancre et on applique la taille SANS déformation
            #    (change "T30" pour décaler : Q/R/S = gauche/droite ; 28/32 = haut/bas)
            anchor_cell = "T30"

            xl_img = XLImage(str(img_file))
            xl_img.width  = out_w
            xl_img.height = out_h
            ws.add_image(xl_img, anchor_cell)
        # sinon on ignore en silence
    except Exception:
        # ne bloque jamais l'export XLSX si l'image plante
        pass

        # --- Logos à gauche du titre (Symbiose + NIKO) ---
    root = _project_root()
    
    def _first_existing(paths):
        for p in paths:
            if p.exists():
                return p
        return None
    
    symbiose_path = _first_existing([
        root / "assets" / "logo_symbiose.png",
        root / "assets" / "signature" / "logo_symbiose.png",
        root / "assets" / "Logo_Symbiose.png",
    ])
    niko_path = _first_existing([
        root / "assets" / "NIKO_Logo.png",
        root / "assets" / "niko_logo.png",
        root / "assets" / "Niko_Logo.png",
    ])
    
    # Positionnement comme le modèle : Symbiose en B3, NIKO juste à côté (E3)
    _add_logo(ws, symbiose_path, anchor_cell="B2", max_w=160, max_h=48)
    _add_logo(ws, niko_path,     anchor_cell="E2", max_w=120, max_h=40)


    # --- H8 : goût (libellé Excel)
    _set(ws, "H8", _to_excel_label(gout1) or "")

    # --- A20 : date de fermentation (semaine_du)
    try:
        _set(ws, "A20", semaine_du, number_format="DD/MM/YYYY")
    except Exception:
        try:
            _set(ws, "A20", date.fromisoformat(str(semaine_du)), number_format="DD/MM/YYYY")
        except Exception:
            _set(ws, "A20", str(semaine_du))

    # --- DDM : garder A10:C10 fusionné ("DDM :") + D10:G10 fusionné (date) ---
    try:
        from openpyxl.styles import Alignment, Font
    
        # 1) Nettoyer les fusions qui chevauchent la zone A10:G10
        for rng in list(ws.merged_cells.ranges):
            if not (rng.max_row < 10 or rng.min_row > 10 or rng.max_col < 1 or rng.min_col > 7):
                ws.unmerge_cells(rng.coord)
    
        # 2) Re-fusionner les zones souhaitées
        ws.merge_cells("A10:C10")
        ws.merge_cells("D10:G10")
    
        # 3) Écrire le libellé et la date
        _safe_set_cell(ws, 10, 1, "DDM :")  # A10 (ancre de A10:C10)
        ws["A10"].alignment = Alignment(vertical="center", horizontal="left")
        try:
            ws["A10"].font = Font(bold=True)
        except Exception:
            pass
    
        _safe_set_cell(ws, 10, 4, ddm, number_format="DD/MM/YYYY")  # D10 (ancre de D10:G10)
        ws["D10"].alignment = Alignment(vertical="center", horizontal="left")  # ou 'right' si tu préfères
    except Exception:
        # Fallback robuste si jamais
        ws.merge_cells("A10:C10")
        ws.merge_cells("D10:G10")
        _safe_set_cell(ws, 10, 1, "DDM :")
        _safe_set_cell(ws, 10, 4, ddm, number_format="DD/MM/YYYY")

    # --- Quantités de cartons par format -> ligne 15 (K/M/O/Q/S/U) ---
    # Mapping voulu :
    # - 33 cL EN (Water kefir)      -> K15 (K:L)
    # - 33 cL FR SANS NIKO          -> M15 (M:N)
    # - 33 cL FR AVEC NIKO          -> O15 (O:P)
    # - 75 cL x6 SANS NIKO          -> Q15 (Q:R)
    # - 75 cL x4 SANS NIKO          -> S15 (S:T)
    # - 75 cL x6 AVEC NIKO          -> U15 (U:V)
    
    k_33_en = 0
    m_33_fr_no = 0
    o_33_fr_niko = 0
    q_75x6 = 0
    s_75x4 = 0
    u_75x6_niko = 0
    
    if isinstance(df_min, pd.DataFrame) and not df_min.empty:
        # Filtre sur le goût choisi (colonne GoutCanon du tableau affiché)
        dff = df_min.copy()
        if "GoutCanon" in dff.columns:
            dff = dff[dff["GoutCanon"].astype(str).str.strip() == str(gout1 or "").strip()]
    
        # Colonne "Cartons à produire ..."
        col_cart = next((c for c in dff.columns if "Cartons à produire" in str(c)), None)
    
        if col_cart and not dff.empty:
            for _, r0 in dff.iterrows():
                qty = pd.to_numeric(r0.get(col_cart), errors="coerce")
                qty = int(qty) if pd.notna(qty) else 0
                if qty <= 0:
                    continue
    
                prod = str(r0.get("Produit", "")).upper()
                stock = str(r0.get("Stock", ""))
    
                nb, volL = _parse_format_from_stock(stock)
                if nb is None or volL is None:
                    # fallback si "Stock" est atypique : on essaie depuis 'Produit'
                    m_nb = re.search(r"x\s*(\d+)", prod)
                    m_vol = re.search(r"(\d+(?:[.,]\d+)?)\s*cL", prod, flags=re.I)
                    nb = int(m_nb.group(1)) if m_nb else nb
                    volL = (float(m_vol.group(1).replace(",", ".")) / 100.0) if m_vol else volL
    
                has_niko = "NIKO" in prod
                is_33cl = (volL is not None) and abs(volL - 0.33) < 1e-6 and nb == 12
                is_75cl = (volL is not None) and abs(volL - 0.75) < 1e-6
                is_english = "WATER KEFIR" in prod  # distingue EN vs FR sur tes libellés
    
                if is_33cl:
                    if is_english:
                        k_33_en += qty            # -> K15
                    else:
                        if has_niko:
                            o_33_fr_niko += qty   # -> O15
                        else:
                            m_33_fr_no += qty     # -> M15
                elif is_75cl:
                    if nb == 6:
                        if has_niko:
                            u_75x6_niko += qty    # -> U15
                        else:
                            q_75x6 += qty         # -> Q15
                    elif nb == 4 and not has_niko:
                        s_75x4 += qty             # -> S15
    
    # Écritures (ligne 15 ; on vise l’ancre des fusions)
    _safe_set_cell(ws, 15, 11, int(k_33_en))      # K15 (K:L)
    _safe_set_cell(ws, 15, 13, int(m_33_fr_no))   # M15 (M:N)
    _safe_set_cell(ws, 15, 15, int(o_33_fr_niko)) # O15 (O:P)
    _safe_set_cell(ws, 15, 17, int(q_75x6))       # Q15 (Q:R)
    _safe_set_cell(ws, 15, 19, int(s_75x4))       # S15 (S:T)
    _safe_set_cell(ws, 15, 21, int(u_75x6_niko))  # U15 (U:V)

    
    # --- Saut de page : forcer "Date de la production" en tête de page suivante ---
    try:
        import re
        from openpyxl.worksheet.pagebreak import Break, PageBreak
    
        def _find_row(ws, pattern: str):
            rx = re.compile(pattern, flags=re.I)
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and rx.search(v):
                        return cell.row
            return None
    
        # repère la ligne de "Date de la production"
        row_date_prod = _find_row(ws, r"\bdate\s+de\s+la\s+production\b")
        if row_date_prod:
            # on vide d'éventuels anciens sauts puis on crée un saut juste avant
            ws.row_breaks = PageBreak()
            ws.row_breaks.append(Break(id=row_date_prod - 1))
    except Exception as e:
        print(f"[xlsx_fill] Saut de page manuel non appliqué: {e}")

    # Sauvegarde en mémoire
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ======================================================================
#                   Remplissage BL enlèvements Sofripa
# ======================================================================

def _iter_cells(ws):
    for r in ws.iter_rows(values_only=False):
        for c in r:
            yield c

def _find_cell_by_regex(ws, pattern: str) -> Tuple[int, int] | Tuple[None, None]:
    rx = re.compile(pattern, flags=re.I)
    for cell in _iter_cells(ws):
        v = cell.value
        if isinstance(v, str) and rx.search(v):
            return cell.row, cell.column
    return None, None

def _write_right_of(ws, row: int, col: int, value):
    """Écrit dans la cellule immédiatement à droite (gère fusions)."""
    _safe_set_cell(ws, row, col + 1, value)

def _write_cell(ws, row: int, col: int, value):
    """Écrit (row,col) en gérant les fusions."""
    _safe_set_cell(ws, row, col, value)

def _normalize_header_text(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("’", "'")
    for ch in ["(", ")", ":", ";", ","]:
        s = s.replace(ch, " ")
    s = " ".join(s.split())
    return s

def _first_data_row_after_header(ws, hdr_row: int, cols: List[int]) -> int:
    """
    Si l'en-tête est fusionné sur 2+ lignes, commence à la 1ère ligne
    *après* ces fusions sur n'importe laquelle des colonnes du tableau.
    """
    start = hdr_row + 1
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= hdr_row <= rng.max_row:
            if any(rng.min_col <= c <= rng.max_col for c in cols if c is not None):
                start = max(start, rng.max_row + 1)
    return start

def fill_bl_enlevements_xlsx(
    template_path: str,
    date_creation: date,
    date_ramasse: date,
    destinataire_title: str,
    destinataire_lines: List[str],
    df_lines: pd.DataFrame,
) -> bytes:
    """
    Remplit le modèle XLSX 'LOG_EN_001_01 BL enlèvements Sofripa-2.xlsx'
    de façon ANCRÉE sur la rangée d'en-têtes réelle (séquence contiguë).
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Modèle Excel introuvable: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ---------- utilitaires locaux ----------
    from openpyxl.styles import Alignment

    def _safe_write(ws, row: int, col: int, value):
        """Écrit en visant l'ancre si la cible est dans une fusion (évite MergedCell)."""
        r, c = row, col
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                r, c = rng.min_row, rng.min_col
                break
        ws.cell(row=r, column=c).value = value

    def _norm(s: str) -> str:
        s = str(s or "").strip().lower()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = s.replace("’", "'")
        for ch in ("(", ")", ":", ";", ","):
            s = s.replace(ch, " ")
        return " ".join(s.split())

    def _find_header_run(ws):
        """
        Trouve la **séquence contiguë et complète** des 6 en-têtes
        et renvoie (row, [c_ref,c_prod,c_ddm,c_qc,c_qp,c_poids]).
        On garde la **plus basse** de la feuille.
        """
        SEQ = [
            ["référence", "reference"],
            ["produit", "produit (gout + format)", "produit gout format"],
            ["ddm", "date de durabilite", "date de durabilité"],
            ["quantité cartons", "quantite cartons", "n° cartons", "no cartons", "nb cartons"],
            ["quantité palettes", "quantite palettes", "n° palettes", "no palettes", "nb palettes"],
            ["poids palettes (kg)", "poids palettes", "poids (kg)"],
        ]
        SEQ = [[_norm(x) for x in alts] for alts in SEQ]

        maxr = min(ws.max_row, 300)
        maxc = min(ws.max_column, 120)
        best = None  # (row, cols)

        for r in range(1, maxr + 1):
            # on parcourt les fenêtres contiguës de 6 cellules
            for c0 in range(1, maxc - 6 + 2):
                ok = True
                cols = []
                for k in range(6):
                    hv = _norm(ws.cell(row=r, column=c0 + k).value)
                    if hv not in SEQ[k]:
                        ok = False
                        break
                    cols.append(c0 + k)
                if ok:
                    # on garde la plus basse (si plusieurs zones existent)
                    best = (r, cols)
        return best

    # ---------- 1) Dates ----------
    def _find_cell_by_regex(ws, pattern: str):
        rx = re.compile(pattern, flags=re.I)
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and rx.search(v):
                    return cell.row, cell.column
        return None, None

    r, c = _find_cell_by_regex(ws, r"date\s+de\s+cr[eé]ation")
    if r and c:
        _safe_write(ws, r, c + 1, date_creation.strftime("%d/%m/%Y"))

    r, c = _find_cell_by_regex(ws, r"date\s+de\s+rammasse|date\s+de\s+ramasse")
    if r and c:
        _safe_write(ws, r, c + 1, date_ramasse.strftime("%d/%m/%Y"))

    # ---------- 2) Destinataire (dans l'encadré, multi-lignes) ----------
    r, c = _find_cell_by_regex(ws, r"destinataire")
    if r and c:
        # essaie de réutiliser une fusion à droite du libellé
        target_rng = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col > c:
                if target_rng is None or rng.min_col < target_rng.min_col:
                    target_rng = rng

        if target_rng:
            rr, cc = target_rng.min_row, target_rng.min_col
            rr_end, cc_end = target_rng.max_row, target_rng.max_col
        else:
            # crée une petite fusion 3x6 à droite si rien n'existe
            rr, cc = r, c + 1
            rr_end, cc_end = min(r + 2, ws.max_row), min(c + 6, ws.max_column)
            try:
                ws.merge_cells(start_row=rr, start_column=cc, end_row=rr_end, end_column=cc_end)
            except Exception:
                pass

        text = "\n".join([destinataire_title] + [x for x in (destinataire_lines or []) if str(x).strip()])
        _safe_write(ws, rr, cc, text)
        a = ws.cell(row=rr, column=cc).alignment or Alignment()
        ws.cell(row=rr, column=cc).alignment = Alignment(wrap_text=True, vertical="top", horizontal=a.horizontal or "left")

        # ajuste la hauteur des lignes fusionnées pour afficher l'adresse
        n_lines = max(1, text.count("\n") + 1)
        span = max(1, rr_end - rr + 1)
        per_row = 14 * n_lines / span
        for rset in range(rr, rr_end + 1):
            cur = ws.row_dimensions[rset].height or 0
            ws.row_dimensions[rset].height = max(cur, per_row)

        # nettoie la ligne parasite si elle traîne en dehors de l'encadré
        zr, zc = _find_cell_by_regex(ws, r"zac\s+du\s+haut\s+de\s+wissous")
        if zr and zc and (zr < rr or zr > rr_end or zc < cc or zc > cc_end):
            _safe_write(ws, zr, zc, "")

    # ---------- 3) Localisation **fiable** de la ligne d’en-têtes ----------
    header = _find_header_run(ws)
    if not header:
        raise KeyError("Impossible de localiser la rangée d’en-têtes (séquence complète non trouvée).")
    hdr_row, (c_ref, c_prod, c_ddm, c_qc, c_qp, c_poids) = header

    # ---------- 4) DataFrame d'entrée : normalisation ----------
    df = df_lines.copy()
    if "Produit" not in df.columns and "Produit (goût + format)" in df.columns:
        df = df.rename(columns={"Produit (goût + format)": "Produit"})

    def _to_ddm_val(x):
        if isinstance(x, date):
            return x.strftime("%d/%m/%Y")
        s = str(x or "").strip()
        if not s:
            return ""
        try:
            if "-" in s and len(s.split("-")[0]) == 4:
                return datetime.strptime(s, "%Y-%m-%d").strftime("%d/%m/%Y")
            return datetime.strptime(s, "%d/%m/%Y").strftime("%d/%m/%Y")
        except Exception:
            return s

    def _as_int(v) -> int:
        try:
            f = float(v)
            return int(round(f))
        except Exception:
            return 0

    # ---------- 5) Écriture des lignes (ancrée) ----------
    row = hdr_row + 1
    for _, r in df.iterrows():
        _safe_write(ws, row, c_ref,   str(r.get("Référence", "")))
        _safe_write(ws, row, c_prod,  str(r.get("Produit", "")))
        _safe_write(ws, row, c_ddm,   _to_ddm_val(r.get("DDM", "")))
        _safe_write(ws, row, c_qc,    _as_int(r.get("Quantité cartons", 0)))
        _safe_write(ws, row, c_qp,    _as_int(r.get("Quantité palettes", 0)))
        _safe_write(ws, row, c_poids, _as_int(r.get("Poids palettes (kg)", 0)))
        row += 1

    # ---------- 6) Sauvegarde ----------
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# =======================  PDF BL enlèvements (fpdf2)  =======================

def build_bl_enlevements_pdf(
    date_creation: date,
    date_ramasse: date,
    destinataire_title: str,
    destinataire_lines: List[str],
    df_lines: pd.DataFrame,
    *,
    # ⬇️ par défaut, on pointe vers ton vrai fichier déjà présent dans le repo
    logo_path: str | None = "assets/signature/logo_symbiose.png",
    issuer_name: str = "FERMENT STATION",
    issuer_lines: List[str] | None = None,
    issuer_footer: str | None = "Produits issus de l'Agriculture Biologique certifié par FR-BIO-01",
) -> bytes:
    """PDF BL au look Excel : encadré, tableau gris, totaux. (Helvetica/latin-1)."""
    from fpdf import FPDF

    # ---------- helpers texte latin-1 ----------
    def _latin1_safe(s: str) -> str:
        s = str(s or "")
        repl = {"—": "-", "–": "-", "‒": "-", "’": "'", "‘": "'", "“": '"', "”": '"', "…": "...",
                "\u00A0": " ", "\u202F": " ", "\u2009": " ", "œ": "oe", "Œ": "OE", "€": "EUR"}
        for k, v in repl.items():
            s = s.replace(k, v)
        return s.encode("latin-1", "replace").decode("latin-1")

    def _txt(x) -> str:
        return _latin1_safe(x)

    # ---------- data ----------
    df = df_lines.copy()
    if "Produit" not in df.columns and "Produit (goût + format)" in df.columns:
        df = df.rename(columns={"Produit (goût + format)": "Produit"})

    def _ival(x):
        try:
            return int(round(float(x)))
        except Exception:
            return 0

    # ---------- PDF ----------
    pdf = FPDF("P", "mm", "A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    left, right = 15, 195
    width = right - left

    # ---- Logo + coordonnées expéditeur
    y = 18
    x_text = left
    if logo_path:
        img_bytes = _load_asset_bytes(logo_path)  # ⬅️ robuste: on charge en bytes depuis la racine du projet
        if img_bytes:
            bio = io.BytesIO(img_bytes)
            pdf.image(bio, x=left, y=y - 2, w=28)  # ajuste w si besoin
            x_text = left + 34  # texte à droite du logo

    pdf.set_xy(x_text, y)
    pdf.set_font("Helvetica", "B", 12); pdf.cell(0, 6, _txt(issuer_name), ln=1)
    pdf.set_x(x_text); pdf.set_font("Helvetica", "", 11)
    if issuer_lines is None:
        issuer_lines = [
            "Carré Ivry Bâtiment D2",
            "47 rue Ernest Renan",
            "94200 Ivry-sur-Seine - FRANCE",
            "Tél : 0967504647",
            "Site : https://www.symbiose-kefir.fr",
        ]
    for line in issuer_lines:
        pdf.set_x(x_text); pdf.cell(0, 5, _txt(line), ln=1)
    if issuer_footer:
        pdf.ln(2); pdf.set_x(x_text); pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 4, _txt(issuer_footer), ln=1)
    pdf.ln(2)

    # ---- Encadré "BON DE LIVRAISON"
    x_box, w_box = left, width * 0.70
    w_lbl, w_val = w_box * 0.55, w_box * 0.45
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_xy(x_box, pdf.get_y() + 2)
    pdf.cell(w_box, 8, _txt("BON DE LIVRAISON"), border=1, ln=1)

    pdf.set_font("Helvetica", "", 11)

    def _row_simple(label: str, value: str):
        pdf.set_x(x_box)
        pdf.cell(w_lbl, 8, _txt(label), border=1)
        pdf.cell(w_val, 8, _txt(value), border=1, ln=1, align="R")

    def _row_dest(label: str, title: str, lines: List[str]):
        val_text = "\n".join([title] + (lines or []))
        n_lines = len(pdf.multi_cell(w_val, 6, _txt(val_text), split_only=True)) or 1
        row_h = max(8, 6 * n_lines)
        y0 = pdf.get_y()
        pdf.set_xy(x_box, y0); pdf.cell(w_lbl, row_h, _txt(label), border=1)
        pdf.set_xy(x_box + w_lbl, y0); pdf.multi_cell(w_val, 6, _txt(val_text), border=1)
        pdf.set_xy(x_box, y0 + row_h)

    _row_simple("DATE DE CREATION :", date_creation.strftime("%d/%m/%Y"))
    _row_simple("DATE DE RAMASSE :", date_ramasse.strftime("%d/%m/%Y"))
    _row_dest("DESTINATAIRE :", destinataire_title, destinataire_lines)

    # ---- Tableau
    pdf.ln(6)
    pdf.set_fill_color(230, 230, 230)

    headers = ["Référence", "Produit", "DDM", "Nb cartons", "Nb palettes", "Poids (kg)"]
    widths_base = [30, 66, 26, 24, 22, 12]
    widths = widths_base[:]
    header_h = 8
    line_h = 6

    pdf.set_font("Helvetica", "B", 10)
    margin_mm = 2.5
    min_w = {0: 30.0, 1: 58.0, 2: 26.0, 3: 22.0, 4: 20.0, 5: 18.0}
    extra_needed = 0.0
    for j, h in enumerate(headers):
        if j == 1:
            continue
        need = pdf.get_string_width(_txt(h)) + 2 * margin_mm
        new_w = max(widths[j], need, min_w.get(j, widths[j]))
        extra_needed += max(0.0, new_w - widths_base[j])
        widths[j] = new_w
    widths[1] = max(min_w[1], widths[1] - extra_needed)
    total = sum(widths)
    if total > 180.0:
        overflow = total - 180.0
        take = min(overflow, max(0.0, widths[1] - min_w[1]))
        widths[1] -= take; overflow -= take
        for j in (3, 4, 5, 0, 2):
            if overflow <= 0: break
            free = max(0.0, widths[j] - min_w[j])
            d = min(free, overflow)
            widths[j] -= d; overflow -= d

    # En-tête
    x = left; y = pdf.get_y()
    for h, w in zip(headers, widths):
        pdf.set_xy(x, y); pdf.cell(w, header_h, _txt(h), border=1, align="C", fill=True); x += w
    pdf.set_xy(left, y + header_h)

    # Lignes
    pdf.set_font("Helvetica", "", 10)
    tot_cart = tot_pal = tot_poids = 0

    def _maybe_break(h):
        if pdf.will_page_break(h + header_h):
            pdf.add_page()
            pdf.set_fill_color(230, 230, 230)
            pdf.set_font("Helvetica", "B", 10)
            xh = left; yh = pdf.get_y()
            for hh, ww in zip(headers, widths):
                pdf.set_xy(xh, yh); pdf.cell(ww, header_h, _txt(hh), border=1, align="C", fill=True); xh += ww
            pdf.set_xy(left, yh + header_h)
            pdf.set_font("Helvetica", "", 10)

    for _, r in df.iterrows():
        ref = _txt(r.get("Référence", ""))
        prod = _txt(r.get("Produit", ""))
        ddm = _txt(r.get("DDM", ""))
        qc = _ival(r.get("Nb cartons", r.get("Quantité cartons", 0)));   tot_cart += qc
        qp = _ival(r.get("Nb palettes", r.get("Quantité palettes", 0)));  tot_pal  += qp
        po = _ival(r.get("Poids (kg)",  r.get("Poids palettes (kg)", 0))); tot_poids += po

        prod_lines = pdf.multi_cell(widths[1], line_h, prod, split_only=True)
        row_h = max(line_h, line_h * len(prod_lines))
        _maybe_break(row_h)

        xrow = left; yrow = pdf.get_y()
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[0], row_h, ref, border=1, align="C"); xrow += widths[0]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[1], line_h, prod, border=1, align="L", max_line_height=line_h); xrow += widths[1]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[2], row_h, ddm, border=1, align="C"); xrow += widths[2]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[3], row_h, str(qc), border=1, align="C"); xrow += widths[3]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[4], row_h, str(qp), border=1, align="C"); xrow += widths[4]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[5], row_h, str(po), border=1, align="C")
        pdf.set_xy(left, yrow + row_h)

    # Totaux
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(widths[0] + widths[1] + widths[2], 8, _txt("Totaux"), border=1, align="R")
    pdf.cell(widths[3], 8, _txt(f"{tot_cart:,}".replace(",", " ")), border=1, align="C")
    pdf.cell(widths[4], 8, _txt(f"{tot_pal:,}".replace(",", " ")),  border=1, align="C")
    pdf.cell(widths[5], 8, _txt(f"{tot_poids:,}".replace(",", " ")), border=1, align="C")

    return bytes(pdf.output(dest="S"))
