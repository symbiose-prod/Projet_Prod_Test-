# core/optimizer.py
import io, re
from pathlib import Path
from typing import Optional, List, Tuple
import numpy as np
import pandas as pd

import unicodedata

def _norm_colname(s: str) -> str:
    s = str(s or "")
    s = s.strip().lower()
    # enlève accents
    s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    # remplace tout le reste par des espaces
    import re as _re
    s = _re.sub(r"[^a-z0-9]+", " ", s)
    s = _re.sub(r"\s+", " ", s).strip()
    return s

def _pick_column(df: pd.DataFrame, candidates_norm: list[str]) -> str | None:
    """
    Retourne le vrai nom de colonne du df correspondant à des candidats "normalisés".
    Amélioré : accepte 'produit 1', 'produit_2', etc. + correspondances partielles.
    """
    norm_to_real = {_norm_colname(c): c for c in df.columns}
    norms = list(norm_to_real.keys())

    # 1) match exact (priorité)
    for cand in candidates_norm:
        if cand in norm_to_real:
            return norm_to_real[cand]

    # 2) startswith sur les mots-clés importants (ex: 'produit' → 'produit 1')
    KEY_PREFIXES = ["produit", "designation", "desigation", "des", "libelle", "libelle", "product", "item", "sku"]
    for key in KEY_PREFIXES:
        for n in norms:
            if n.startswith(key):
                return norm_to_real[n]

    # 3) contains (au cas où un préfixe/suffixe se glisse)
    for key in KEY_PREFIXES:
        for n in norms:
            if key in n:
                return norm_to_real[n]

    # 4) fuzzy (secours)
    try:
        import difflib
        match = difflib.get_close_matches(candidates_norm[0], norms, n=1, cutoff=0.85)
        if match:
            return norm_to_real[match[0]]
    except Exception:
        pass
    return None


# ======= constantes
ALLOWED_FORMATS = {(12, 0.33), (6, 0.75), (4, 0.75)}
ROUND_TO_CARTON = True
VOL_TOL = 0.02
EPS = 1e-9
DEFAULT_WINDOW_DAYS = 60

# ---------- Helpers sélection et égalisation ----------

def _weekly_perte(stock_hl: float, vitesse_hl_j: float, price_hL: float = 400.0) -> float:
    """Perte € sur 7 jours si on ne produit pas : max(demande7 - stock, 0) * prix."""
    dem7 = 7.0 * max(float(vitesse_hl_j), 0.0)
    manque = max(dem7 - max(float(stock_hl), 0.0), 0.0)
    return manque * float(price_hL)

def _equalize_last_batch_global(Gi: np.ndarray, vi: np.ndarray, V: float) -> np.ndarray:
    """
    Égalise un horizon unique T sur T = (Gi + xi)/vi pour un groupe donné.
    Résout sum_i max(0, T*vi - Gi) = V par dichotomie (xi >= 0).
    Retourne x (hL) par ligne.
    """
    vi = np.maximum(vi.astype(float), 0.0)
    Gi = np.maximum(Gi.astype(float), 0.0)
    if V <= 1e-12 or vi.sum() <= 1e-12:
        return np.zeros_like(Gi)

    # bornes : T_min = max(Gi/vi) (horizon sans prod), T_max = T_min + marge
    with np.errstate(divide='ignore', invalid='ignore'):
        T0 = np.nanmax(np.where(vi > 0, Gi / np.maximum(vi, 1e-12), 0.0))
    T_lo = T0
    T_hi = T0 + (V / max(np.max(vi), 1e-12)) + 365.0  # marge large

    # dichotomie
    for _ in range(80):
        T_mid = 0.5 * (T_lo + T_hi)
        x = np.maximum(T_mid * vi - Gi, 0.0)
        s = x.sum()
        if s > V:
            T_hi = T_mid
        else:
            T_lo = T_mid
    x = np.maximum(T_lo * vi - Gi, 0.0)

    # petit rescale pour coller à V
    s = x.sum()
    if s > 0:
        x *= (V / s)
    return x


# ======= util accents (fix_text)
ACCENT_CHARS = "éèêëàâäîïôöùûüçÉÈÊËÀÂÄÎÏÔÖÙÛÜÇ"
CUSTOM_REPLACEMENTS = {
    "M�lisse": "Mélisse",
    "poivr�e": "poivrée",
    "P�che": "Pêche",
}
def _looks_better(a: str, b: str) -> bool:
    def score(s): return sum(ch in ACCENT_CHARS for ch in s)
    return score(b) > score(a)
def fix_text(s) -> str:
    if s is None: return ""
    if not isinstance(s, str): s = str(s)
    s0 = s
    try:
        s1 = s0.encode("latin1").decode("utf-8")
        if _looks_better(s0, s1): s0 = s1
    except Exception:
        pass
    if s0 in CUSTOM_REPLACEMENTS: return CUSTOM_REPLACEMENTS[s0]
    if "�" in s0: s0 = s0.replace("�", "é")
    return s0

# ======= détection en-tête & période B2 (sans Streamlit)
def detect_header_row(df_raw: pd.DataFrame) -> int:
    must = {"Produit", "Stock", "Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"}
    for i in range(min(10, len(df_raw))):
        if must.issubset(set(str(x).strip() for x in df_raw.iloc[i].tolist())):
            return i
    return 0

def rows_to_keep_by_fill(excel_bytes: bytes, header_idx: int) -> List[bool]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    start_row = header_idx + 2
    keep: List[bool] = []
    for r in range(start_row, ws.max_row + 1):
        is_black = False
        for cell in ws[r]:
            fill = cell.fill
            if fill and fill.fill_type:
                rgb = getattr(getattr(fill, "fgColor", None), "rgb", None) or getattr(getattr(fill, "start_color", None), "rgb", None)
                if rgb and rgb[-6:].upper() == "000000":
                    is_black = True
                    break
        keep.append(not is_black)
    return keep

def parse_days_from_b2(value) -> Optional[int]:
    try:
        if isinstance(value, (int, float)) and not pd.isna(value):
            v = int(round(float(value)));  return v if v > 0 else None
        if value is None: return None
        s = str(value).strip()
        m = re.search(r"(\d+)\s*(?:j|jour|jours)\b", s, flags=re.IGNORECASE)
        if m: return int(m.group(1)) or None
        date_pat = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}).*?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})"
        m2 = re.search(date_pat, s)
        if m2:
            d1 = pd.to_datetime(m2.group(1), dayfirst=True, errors="coerce")
            d2 = pd.to_datetime(m2.group(2), dayfirst=True, errors="coerce")
            if pd.notna(d1) and pd.notna(d2):
                days = int((d2 - d1).days)
                return days if days > 0 else None
        m3 = re.search(r"\b(\d{1,4})\b", s)
        if m3:
            v = int(m3.group(1));  return v if v > 0 else None
    except Exception:
        return None
    return None

def read_input_excel_and_period_from_path(path_xlsx: str) -> Tuple[pd.DataFrame, int]:
    with open(path_xlsx, "rb") as f:
        file_bytes = f.read()
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    header_idx = detect_header_row(raw)
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    keep_mask = rows_to_keep_by_fill(file_bytes, header_idx)
    if len(keep_mask) < len(df):
        keep_mask = keep_mask + [True] * (len(df) - len(keep_mask))
    df = df.iloc[[i for i, k in enumerate(keep_mask) if k]].reset_index(drop=True)
    # lecture B2
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        b2_val = ws["B2"].value
        wd = parse_days_from_b2(b2_val)
    except Exception:
        wd = None
    return df, (wd if wd and wd > 0 else DEFAULT_WINDOW_DAYS)

# ======= flavor map
def load_flavor_map_from_path(path_csv: str) -> pd.DataFrame:
    import csv
    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin1"]
    seps = [",", ";", "\t", "|"]
    if not Path(path_csv).exists():
        return pd.DataFrame(columns=["name", "canonical"])
    for enc in encodings:
        for sep in seps:
            try:
                fm = pd.read_csv(path_csv, encoding=enc, sep=sep, engine="python")
                lower = {c.lower(): c for c in fm.columns}
                if "name" in lower and "canonical" in lower:
                    fm = fm[[lower["name"], lower["canonical"]]].copy()
                    fm.columns = ["name","canonical"]
                    fm = fm.dropna()
                    fm["name"] = fm["name"].astype(str).str.strip().map(fix_text)
                    fm["canonical"] = fm["canonical"].astype(str).str.strip().map(fix_text)
                    fm = fm[(fm["name"]!="") & (fm["canonical"]!="")]
                    return fm
            except Exception:
                continue
    return pd.DataFrame(columns=["name", "canonical"])

def apply_canonical_flavor(df: pd.DataFrame, fm: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    # 1) Trouve la colonne "Produit" même si le nom diffère (Désignation, Libellé, Product, etc.)
    prod_candidates = [
        "produit", "produit 1", "produit1", "produit 2",
        "designation", "désignation", "libelle", "libellé",
        "nom du produit", "product", "sku libelle", "sku libellé", "sku", "item"
    ]
    prod_candidates = [_norm_colname(x) for x in prod_candidates]
    col_prod = _pick_column(out, prod_candidates)

    if not col_prod:
        # message clair si rien n'est trouvé
        cols_list = ", ".join(map(str, out.columns))
        raise KeyError(
            "Colonne produit introuvable. "
            "Renomme la colonne en 'Produit' ou 'Désignation' (ou équivalent). "
            f"Colonnes détectées: {cols_list}"
        )

    # 2) Crée la colonne standard 'Produit'
    out["Produit"] = out[col_prod].astype(str).map(fix_text)
    out["Produit_norm"] = out["Produit"].str.strip()

    # 3) Mapping canonique
    if len(fm):
        fm = fm.dropna(subset=["name","canonical"]).copy()
        fm["name_norm"] = fm["name"].astype(str).map(fix_text).str.strip().str.lower()
        fm["canonical"] = fm["canonical"].astype(str).map(fix_text).str.strip()
        m_exact = dict(zip(fm["name_norm"], fm["canonical"]))
        keys = list(m_exact.keys())
        import difflib as _difflib
        def to_canonical(prod: str) -> str:
            s = str(prod).strip().lower()
            if s in m_exact: return m_exact[s]
            try:
                close = _difflib.get_close_matches(s, keys, n=1, cutoff=0.92)
                if close: return m_exact[close[0]]
            except Exception:
                pass
            return str(prod).strip()
        out["GoutCanon"] = out["Produit_norm"].map(to_canonical)
    else:
        out["GoutCanon"] = out["Produit_norm"]

    out["GoutCanon"] = out["GoutCanon"].astype(str).map(fix_text).str.strip()
    return out


# ======= parsing formats/stock & filtres
def parse_stock(text: str):
    if pd.isna(text): return np.nan, np.nan
    s = str(text)
    nb = None
    for pat in [r"(?:Carton|Caisse|Colis)\s+de\s*(\d+)", r"(\d+)\s*[x×]\s*Bouteilles?", r"(\d+)\s*[x×]", r"(\d+)\s+Bouteilles?"]:
        m = re.search(pat, s, flags=re.IGNORECASE)
        if m:
            try: nb = int(m.group(1)); break
            except: pass
    vol_l = None
    m_l = re.findall(r"(\d+(?:[.,]\d+)?)\s*[lL]", s)
    if m_l: vol_l = float(m_l[-1].replace(",", "."))
    else:
        m_cl = re.findall(r"(\d+(?:[.,]\d+)?)\s*c[lL]", s)
        if m_cl: vol_l = float(m_cl[-1].replace(",", ".")) / 100.0
    if nb is None or vol_l is None:
        m_combo = re.search(r"(\d+)\s*[x×]\s*(\d+(?:[.,]\d+)?)+\s*([lc]l?)", s, flags=re.IGNORECASE)
        if m_combo:
            try:
                nb2 = int(m_combo.group(1)); val = float(m_combo.group(2).replace(",", "."))
                unit = m_combo.group(3).lower(); vol2 = val if unit.startswith("l") else val/100.0
                if nb is None: nb = nb2
                if vol_l is None: vol_l = vol2
            except: pass
    if (nb is None or np.isnan(nb)) and vol_l is not None and abs(vol_l - 0.75) <= VOL_TOL:
        if re.search(r"(?:\b4\s*[x×]\b|Carton\s+de\s*4\b|4\s+Bouteilles?)", s, flags=re.IGNORECASE):
            nb = 4
    return (float(nb) if nb is not None else np.nan, float(vol_l) if vol_l is not None else np.nan)

def safe_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")

def is_allowed_format(nb_bottles, vol_l, stock_txt: str) -> bool:
    if pd.isna(nb_bottles) or pd.isna(vol_l):
        if re.search(r"(?:\b4\s*[x×]\s*75\s*c?l\b|\b4\s+Bouteilles?\b.*75\s*c?l)", stock_txt, flags=re.IGNORECASE):
            nb_bottles = 4; vol_l = 0.75
        else:
            return False
    nb_bottles = int(nb_bottles); vol_l = float(vol_l)
    for nb_ok, vol_ok in ALLOWED_FORMATS:
        if nb_bottles == nb_ok and abs(vol_l - vol_ok) <= VOL_TOL:
            return True
    return False

BLOCKED_LABELS_EXACT = {"Autres (coffrets, goodies...)"}
BLOCKED_LABELS_LOWER = {"nan", "none", ""}

def sanitize_gouts(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["GoutCanon"] = out["GoutCanon"].astype(str).str.strip()
    mask = ~out["GoutCanon"].str.lower().isin(BLOCKED_LABELS_LOWER)
    mask &= ~out["GoutCanon"].isin(BLOCKED_LABELS_EXACT)
    return out.loc[mask].reset_index(drop=True)

# ======= calculs principaux
def compute_plan(df_in, window_days, volume_cible, nb_gouts, repartir_pro_rv, manual_keep, exclude_list):
    required = ["Produit", "GoutCanon", "Stock", "Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"]
    miss = [c for c in required if c not in df_in.columns]
    if miss: raise ValueError(f"Colonnes manquantes: {miss}")

    # --- helper catégorie ---
    def _category(g: str) -> str:
        s = str(g or "").strip().lower()
        return "infusion" if "infusion" in s else "kefir"

    note_msg = ""  # message d’ajustement à renvoyer à l’UI

    df = df_in[required].copy()
    for c in ["Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"]:
        df[c] = safe_num(df[c])

    parsed = df["Stock"].apply(parse_stock)
    df[["Bouteilles/carton", "Volume bouteille (L)"]] = pd.DataFrame(parsed.tolist(), index=df.index)
    mask_allowed = df.apply(lambda r: is_allowed_format(r["Bouteilles/carton"], r["Volume bouteille (L)"], str(r["Stock"])), axis=1)
    df = df.loc[mask_allowed].reset_index(drop=True)

    df["Volume/carton (hL)"] = (df["Bouteilles/carton"] * df["Volume bouteille (L)"]) / 100.0
    df = df.dropna(subset=["GoutCanon", "Volume/carton (hL)", "Volume vendu (hl)", "Volume disponible (hl)"]).reset_index(drop=True)

    df_all_formats = df.copy()

    if exclude_list:
        ex = {s.strip() for s in exclude_list}
        df = df[~df["GoutCanon"].astype(str).str.strip().isin(ex)]

    if manual_keep:
        keep = {g.strip() for g in manual_keep}
        df = df[df["GoutCanon"].astype(str).str.strip().isin(keep)]

    # --- agrégats par goût ---
    agg = df.groupby("GoutCanon").agg(
        ventes_hl=("Volume vendu (hl)", "sum"),
        stock_hl=("Volume disponible (hl)", "sum")
    )

    # --- Sélection : rupture -> perte € -> autonomie ---
    agg["vitesse_j"] = agg["ventes_hl"] / max(float(window_days), 1.0)
    dem7 = 7.0 * agg["vitesse_j"]
    agg["rupture_semaine"] = agg["stock_hl"] < dem7 - 1e-9
    PRICE_REF = 400.0
    agg["perte_7j"] = np.maximum(dem7 - agg["stock_hl"], 0.0) * PRICE_REF
    agg["autonomie_j"] = np.where(agg["vitesse_j"] > 0, agg["stock_hl"] / agg["vitesse_j"], np.inf)

    agg = agg.sort_values(
        by=["rupture_semaine", "perte_7j", "autonomie_j"],
        ascending=[False, False, True]
    )

    if not manual_keep:
        g_rupt  = [g for g, r in zip(agg.index.tolist(), agg["rupture_semaine"].tolist()) if r]
        g_other = [g for g, r in zip(agg.index.tolist(), agg["rupture_semaine"].tolist()) if not r]
        gouts_cibles = (g_rupt + g_other)[:nb_gouts]
    else:
        gouts_cibles = sorted(set(df["GoutCanon"]))
        if len(gouts_cibles) > nb_gouts:
            order = [g for g in agg.index if g in gouts_cibles]
            gouts_cibles = order[:nb_gouts]

    # --- Contrainte dure : si 2 goûts → même catégorie (Infusion OU Kéfir) ---
    if nb_gouts == 2 and len(gouts_cibles) == 2:
        def _rank_candidates_for_category(cat: str) -> list[str]:
            """Tri interne: rupture (desc) → perte € (desc) → autonomie (asc)."""
            pool = [g for g in agg.index if _category(g) == cat]
            if not pool:
                return []
            sub = agg.loc[pool, ["rupture_semaine", "perte_7j", "autonomie_j"]].copy()
            sub["__key__"] = list(sub.index)
            sub = sub.sort_values(
                by=["rupture_semaine", "perte_7j", "autonomie_j"],
                ascending=[False, False, True]
            )
            return sub["__key__"].tolist()

        cats = ["infusion", "kefir"]
        ranked_by_cat = {c: _rank_candidates_for_category(c) for c in cats}
        valid = [c for c in cats if len(ranked_by_cat[c]) >= 2]

        if valid:
            if len(valid) == 1:
                choose = valid[0]
            else:
                order_global = list(agg.index)
                pos = {c: order_global.index(ranked_by_cat[c][0]) for c in valid}
                choose = min(valid, key=lambda c: pos[c])  # meilleur candidat global

            new_pair = ranked_by_cat[choose][:2]
            if set(new_pair) != set(gouts_cibles):
                note_msg = (
                    "⚠️ Contrainte appliquée : pas de co-production **Infusion + Kéfir**. "
                    f"Sélection ajustée → deux recettes **{ 'Infusion' if choose=='infusion' else 'Kéfir' }** "
                    f"({new_pair[0]} ; {new_pair[1]})."
                )
            gouts_cibles = new_pair
        # sinon : pas 2 goûts dispo dans une même catégorie → on garde tel quel

    df_selected = df[df["GoutCanon"].isin(gouts_cibles)].copy()
    if len(gouts_cibles) == 0:
        raise ValueError("Aucun goût sélectionné.")

    # ---- ALLOCATION PAR GOÛT : égalise le jour d'épuisement entre formats d'un même goût ----
    df_calc = df_selected.copy()
    df_calc["v_i"] = df_calc["Volume vendu (hl)"] / max(float(window_days), 1.0)
    df_calc["G_i (hL)"] = df_calc["Volume disponible (hl)"]
    V_tot = float(volume_cible)

    # Partage du volume entre goûts (prorata ventes si dispo, sinon égalitaire)
    ventes_par_gout = df_calc.groupby("GoutCanon")["Volume vendu (hl)"].sum()
    pos = ventes_par_gout > 0
    if repartir_pro_rv and pos.any():
        w_gout = (ventes_par_gout[pos] / ventes_par_gout[pos].sum()).reindex(ventes_par_gout.index, fill_value=0.0)
    else:
        n = max(len(ventes_par_gout), 1)
        w_gout = pd.Series(1.0 / n, index=ventes_par_gout.index)

    df_calc["X_adj (hL)"] = 0.0
    for g, grp in df_calc.groupby("GoutCanon"):
        Vg = V_tot * float(w_gout.get(g, 0.0))
        Gi = grp["G_i (hL)"].to_numpy(float)
        vi = np.maximum(grp["v_i"].to_numpy(float), 0.0)
        xg = _equalize_last_batch_global(Gi, vi, Vg)  # égalise (Gi+x_i)/v_i = T_g (constant dans le goût)
        df_calc.loc[grp.index, "X_adj (hL)"] = np.maximum(xg, 0.0)

    cap_resume = f"{volume_cible:.2f} hL au total (égalité du jour d'épuisement par goût)"

    # ---- conversions cartons / bouteilles ----
    df_calc["Cartons à produire (exact)"] = df_calc["X_adj (hL)"] / df_calc["Volume/carton (hL)"]
    if ROUND_TO_CARTON:
        df_calc["Cartons à produire (arrondi)"] = np.floor(df_calc["Cartons à produire (exact)"] + 0.5).astype("Int64")
        df_calc["Volume produit arrondi (hL)"] = df_calc["Cartons à produire (arrondi)"] * df_calc["Volume/carton (hL)"]

    df_calc["Bouteilles à produire (exact)"] = df_calc["Cartons à produire (exact)"] * df_calc["Bouteilles/carton"]
    if ROUND_TO_CARTON:
        df_calc["Bouteilles à produire (arrondi)"] = (
            df_calc["Cartons à produire (arrondi)"] * df_calc["Bouteilles/carton"]
        ).astype("Int64")

    df_min = df_calc[[
        "GoutCanon", "Produit", "Stock",
        "Cartons à produire (arrondi)",
        "Bouteilles à produire (arrondi)",
        "Volume produit arrondi (hL)"
    ]].sort_values(["GoutCanon", "Produit", "Stock"]).reset_index(drop=True)

    # ---- synthèse sélection ----
    agg_full = df.groupby("GoutCanon").agg(
        ventes_hl=("Volume vendu (hl)", "sum"),
        stock_hl=("Volume disponible (hl)", "sum")
    )
    agg_full["vitesse_j"] = agg_full["ventes_hl"] / max(float(window_days), 1.0)
    agg_full["jours_autonomie"] = np.where(agg_full["vitesse_j"] > 0, agg_full["stock_hl"] / agg_full["vitesse_j"], np.inf)
    agg_full["score_urgence"] = agg_full["vitesse_j"] / (agg_full["jours_autonomie"] + EPS)
    sel_gouts = sorted(set(df_calc["GoutCanon"]))
    synth_sel = agg_full.loc[sel_gouts][["ventes_hl", "stock_hl", "vitesse_j", "jours_autonomie", "score_urgence"]].copy()
    synth_sel = synth_sel.rename(columns={
        "ventes_hl": "Ventes 2 mois (hL)",
        "stock_hl": "Stock (hL)",
        "vitesse_j": "Vitesse (hL/j)",
        "jours_autonomie": "Autonomie (jours)",
        "score_urgence": "Score urgence"
    })

    # 7 sorties (comme utilisé par la page Production)
    return df_min, cap_resume, sel_gouts, synth_sel, df_calc, df, note_msg


def compute_losses_table_v48(df_in_all: pd.DataFrame, window_days: float, price_hL: float) -> pd.DataFrame:
    out_cols = ["Goût", "Demande 7 j (hL)", "Stock (hL)", "Manque sur 7 j (hL)", "Prix moyen (€/hL)", "Perte (€)"]
    if df_in_all is None or not isinstance(df_in_all, pd.DataFrame) or df_in_all.empty:
        return pd.DataFrame(columns=out_cols)
    df = df_in_all.copy()
    if "GoutCanon" not in df.columns:
        return pd.DataFrame(columns=out_cols)
    for c in ["Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df["GoutCanon"] = df["GoutCanon"].astype(str).str.strip()
    bad_lower = {"nan", "none", ""}
    df = df[~df["GoutCanon"].str.lower().isin(bad_lower)]
    df = df[df["GoutCanon"] != "Autres (coffrets, goodies...)"]
    if df.empty:
        return pd.DataFrame(columns=out_cols)
    jours = max(float(window_days), 1.0)
    agg = df.groupby("GoutCanon", as_index=False).agg(
        ventes_hL=("Volume vendu (hl)", "sum"),
        stock_hL=("Volume disponible (hl)", "sum"),
    )
    if agg.empty:
        return pd.DataFrame(columns=out_cols)
    agg["vitesse_hL_j"] = agg["ventes_hL"] / jours
    agg["Demande 7 j (hL)"] = 7.0 * agg["vitesse_hL_j"]
    agg["Stock (hL)"] = agg["stock_hL"]
    agg["Manque sur 7 j (hL)"] = np.clip(agg["Demande 7 j (hL)"] - agg["Stock (hL)"], a_min=0.0, a_max=None)
    agg["Prix moyen (€/hL)"] = float(price_hL)
    agg["Perte (€)"] = (agg["Manque sur 7 j (hL)"] * agg["Prix moyen (€/hL)"]).round(0)
    pertes = agg.rename(columns={"GoutCanon": "Goût"})[
        ["Goût", "Demande 7 j (hL)", "Stock (hL)", "Manque sur 7 j (hL)", "Prix moyen (€/hL)", "Perte (€)"]
    ]
    pertes["Goût"] = pertes["Goût"].map(fix_text)
    pertes["Demande 7 j (hL)"] = pertes["Demande 7 j (hL)"].round(2)
    pertes["Stock (hL)"] = pertes["Stock (hL)"].round(2)
    pertes["Manque sur 7 j (hL)"] = pertes["Manque sur 7 j (hL)"].round(2)
    pertes["Prix moyen (€/hL)"] = pertes["Prix moyen (€/hL)"].round(0)
    return pertes.sort_values("Perte (€)", ascending=False).reset_index(drop=True)

# --- LECTURE EXCEL depuis un UPLOAD Streamlit (sans rien changer ailleurs) ---

def read_input_excel_and_period_from_bytes(file_bytes: bytes):
    """Même logique que _from_path mais pour des bytes (uploader Streamlit)."""
    import io, openpyxl
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    header_idx = detect_header_row(raw)
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    keep_mask = rows_to_keep_by_fill(file_bytes, header_idx)
    if len(keep_mask) < len(df):
        keep_mask = keep_mask + [True] * (len(df) - len(keep_mask))
    df = df.iloc[[i for i, k in enumerate(keep_mask) if k]].reset_index(drop=True)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        b2_val = ws["B2"].value
        wd = parse_days_from_b2(b2_val)
    except Exception:
        wd = None
    return df, (wd if wd and wd > 0 else DEFAULT_WINDOW_DAYS)

def read_input_excel_and_period_from_upload(uploaded_file):
    """Wrapper pratique pour st.file_uploader (obj upload Streamlit)."""
    file_bytes = uploaded_file.read()
    return read_input_excel_and_period_from_bytes(file_bytes)
