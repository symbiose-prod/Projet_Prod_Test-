# pages/04_Achats_conditionnements.py
from __future__ import annotations
import io, re, unicodedata
from typing import Tuple, List, Dict

import numpy as np
import pandas as pd
import streamlit as st

from common.design import apply_theme, section, kpi
from core.optimizer import parse_stock, VOL_TOL  # formats 12x33 / 6x75 / 4x75


# ====================== UI (entête) ======================
apply_theme("Achats — Conditionnements", "📦")
section("Prévision d’achats (conditionnements)", "📦")

# Besoin du fichier ventes déjà chargé dans l'accueil
if "df_raw" not in st.session_state or "window_days" not in st.session_state:
    st.warning("Va d’abord dans **Accueil** pour déposer l’Excel des ventes/stock, puis reviens ici.")
    st.stop()

df_raw = st.session_state.df_raw.copy()
window_days = float(st.session_state.window_days)

# ---------------- Sidebar (période + options) ----------------
with st.sidebar:
    st.header("Période à prévoir")
    horizon_j = st.number_input("Horizon (jours)", min_value=1, max_value=365, value=14, step=1)
    st.caption("Le besoin prévoit une consommation sur cet horizon à partir des ventes moyennes.")
    st.markdown("---")
    st.header("Options étiquettes")
    force_labels = st.checkbox(
        "Étiquettes = 1 par bouteille (forcer si 'étiquette' dans le nom)",
        value=True
    )

st.caption(
    f"Excel ventes courant : **{st.session_state.get('file_name','(sans nom)')}** — "
    f"Fenêtre de calcul des vitesses : **{int(window_days)} jours** — "
    f"Horizon prévision : **{int(horizon_j)} jours**"
)

# ====================== IMPORTS (dans la page) ======================
section("Importer les fichiers", "📥")
c1, c2 = st.columns(2)
with c1:
    st.subheader("Consommation des articles (Excel)")
    conso_file = st.file_uploader(
        "Déposer le fichier *Consommation* ici",
        type=["xlsx","xls"],
        key="uploader_conso",
        label_visibility="collapsed"
    )
with c2:
    st.subheader("Stocks des articles (Excel)")
    stock_file = st.file_uploader(
        "Déposer le fichier *Stocks* ici",
        type=["xlsx","xls"],
        key="uploader_stock",
        label_visibility="collapsed"
    )

# ====================== Helpers généraux ======================

def _norm_txt(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s

def _canon_txt(s: str) -> str:
    s = str(s or "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-zA-Z0-9]+", " ", s).strip().lower()
    return s

def _is_total_row(s: str) -> bool:
    """True si libellé est une ligne de total (TOTAL, Total général, …)."""
    t = _canon_txt(s)
    if not t:
        return False
    if t.startswith("total"):
        return True
    return t in {
        "total general", "grand total", "totaux", "total stock",
        "total stocks", "total consommation", "total consommations",
        "total achats", "total des achats"
    }

def _find_cell(df_nohdr: pd.DataFrame, pattern: str) -> Tuple[int | None, int | None]:
    pat = _norm_txt(pattern)
    for r in range(df_nohdr.shape[0]):
        row = df_nohdr.iloc[r].astype(str).tolist()
        for c, v in enumerate(row):
            if pat in _norm_txt(v):
                return r, c
    return None, None

def _parse_number(x: str | float | int) -> float:
    """Tolère , décimales et séparateurs d'espace/point pour milliers."""
    if isinstance(x, (int, float)) and not pd.isna(x):
        return float(x)
    s = str(x or "").strip()
    if not s:
        return np.nan
    s = s.replace("\u202f", " ").replace("\xa0", " ")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(" ", "")
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return np.nan

def _parse_days_from_b2(value) -> int | None:
    """
    Accepte:
      - un entier (jours)
      - une chaîne "xx jours"
      - une plage de dates "01/08/2025 au 31/08/2025" -> (d2-d1).days
    """
    try:
        if isinstance(value, (int, float)) and not pd.isna(value):
            v = int(round(float(value)));  return v if v > 0 else None
        if value is None: return None
        s = str(value).strip()
        m = re.search(r"(\d+)\s*(?:j|jour|jours)\b", s, flags=re.IGNORECASE)
        if m: return max(int(m.group(1)), 1)
        date_pat = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}).*?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})"
        m2 = re.search(date_pat, s)
        if m2:
            d1 = pd.to_datetime(m2.group(1), dayfirst=True, errors="coerce")
            d2 = pd.to_datetime(m2.group(2), dayfirst=True, errors="coerce")
            if pd.notna(d1) and pd.notna(d2):
                days = int((d2 - d1).days)
                return days if days > 0 else None
        m3 = re.search(r"\b(\d{1,4})\b", s)
        if m3:
            v = int(m3.group(1));  return v if v > 0 else None
    except Exception:
        return None
    return None

@st.cache_data(show_spinner=False)
def read_consumption_xlsx(file) -> Tuple[pd.DataFrame, int]:
    """
    Extrait la zone :
      - colonne ARTICLE = la colonne où se trouve le mot 'conditionnement'
      - colonne CONSO   = la colonne immédiatement à droite (ou la 1re numérique à droite)
    Lignes : à partir de la ligne sous 'conditionnement' et jusqu'à **2 lignes avant**
    la ligne qui contient 'contenants'. Ignore les lignes 'TOTAL'.
    Retourne (df, conso_days) avec df = colonnes [key, article, conso, per_hint].
    """
    # On lit d'abord B2 via openpyxl
    try:
        import openpyxl
        b = file.read() if hasattr(file, "read") else file
        if isinstance(b, (bytes, bytearray)):
            bio = io.BytesIO(b)
        else:
            file.seek(0); bio = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(bio, data_only=True)
        ws = wb[wb.sheetnames[0]]
        b2_val = ws["B2"].value
        conso_days = _parse_days_from_b2(b2_val) or 30
        bio.seek(0)
        df0 = pd.read_excel(bio, header=None, dtype=str)
    except Exception:
        conso_days = 30
        file.seek(0)
        df0 = pd.read_excel(file, header=None, dtype=str)

    # util local
    def _norm_txt_local(s: str) -> str:
        s = str(s or "").strip().lower()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = re.sub(r"\s+", " ", s)
        return s

    # Trouver meilleure ancre "conditionnement"
    anchors = []
    for r in range(df0.shape[0]):
        for c in range(df0.shape[1]):
            if "conditionnement" in _norm_txt_local(df0.iat[r, c]):
                k = 0
                rr = r + 1
                while rr < df0.shape[0] and str(df0.iat[rr, c]).strip():
                    k += 1; rr += 1
                anchors.append((k, r, c))
    if not anchors:
        raise RuntimeError("Mot-clé 'conditionnement' introuvable dans le fichier consommation.")
    _, r_cond, c_cond = max(anchors)

    # Limite haute : 2 lignes avant la 1re occurrence de "contenants" sous l'ancre
    r_stop = None
    for r in range(r_cond + 1, df0.shape[0]):
        row_txt = " ".join(str(x) for x in df0.iloc[r].tolist())
        if "contenants" in _norm_txt_local(row_txt):
            r_stop = r; break
    if r_stop is None: r_stop = df0.shape[0]

    row_start = r_cond + 1
    row_end   = max(row_start, r_stop - 2)

    # Choix colonnes: article = colonne ancre ; conso = colonne numérique à droite (priorité c_cond+1)
    def _count_numeric(col_idx: int) -> int:
        vals = df0.iloc[row_start:row_end, col_idx].astype(str)
        vals = vals.str.replace(",", ".", regex=False)
        x = pd.to_numeric(vals, errors="coerce")
        return int(x.notna().sum())

    col_article = c_cond
    col_val = c_cond + 1
    if col_val >= df0.shape[1] or _count_numeric(col_val) == 0:
        best = None
        for cc in range(c_cond + 1, df0.shape[1]):
            cnt = _count_numeric(cc)
            if cnt > 0:
                best = (cnt, cc) if best is None or cnt > best[0] else best
        if best is None:
            raise RuntimeError("Impossible de trouver la colonne de **consommation** numérique à droite.")
        col_val = best[1]

    block = df0.iloc[row_start:row_end, [col_article, col_val]].copy()
    block.columns = ["article", "conso_raw"]
    block["article"] = block["article"].astype(str).str.strip()
    block = block[block["article"].map(lambda s: not _is_total_row(s))]
    block["conso"] = pd.to_numeric(block["conso_raw"].astype(str).str.replace(",", ".", regex=False),
                                   errors="coerce").fillna(0.0)

    # Heuristique unité
    def _per_hint(a: str) -> str:
        a0 = _norm_txt_local(a)
        return "carton" if any(w in a0 for w in ["carton", "caisse", "colis", "etui", "étui"]) else "bottle"

    block["per_hint"] = block["article"].map(_per_hint)
    block["key"] = block["article"].map(_norm_txt_local)
    block = block.groupby(["key", "article", "per_hint"], as_index=False)["conso"].sum()

    return block[["key", "article", "conso", "per_hint"]], int(conso_days)

@st.cache_data(show_spinner=False)
def read_stock_xlsx(file) -> pd.DataFrame:
    """Repère l'en-tête 'Quantité virtuelle' et lit les stocks (en filtrant les TOTAL)."""
    df0 = pd.read_excel(file, header=None, dtype=str)
    r_hdr, c_q = _find_cell(df0, "quantité virtuelle")
    if r_hdr is None:
        raise RuntimeError("En-tête 'Quantité virtuelle' introuvable dans l'Excel de stocks.")

    name_candidates = {"article", "designation", "désignation", "libelle", "libellé"}
    c_name = None
    for cc in range(df0.shape[1]):
        if _norm_txt(str(df0.iloc[r_hdr, cc])) in name_candidates:
            c_name = cc; break
    if c_name is None:
        for cc in range(max(0, c_q - 1), -1, -1):
            if str(df0.iloc[r_hdr, cc]).strip():
                c_name = cc; break
    if c_name is None: c_name = 0

    body = df0.iloc[r_hdr + 1 :, [c_name, c_q]].copy()
    body.columns = ["article", "stock_raw"]
    body["article"] = body["article"].astype(str).str.strip()
    body = body[body["article"].str.len() > 0]
    body = body[~body["article"].map(_is_total_row)]

    body["stock"] = pd.to_numeric(body["stock_raw"].map(_parse_number), errors="coerce").fillna(0.0)
    body["key"] = body["article"].map(_norm_txt)
    body = body.groupby(["key", "article"], as_index=False)["stock"].sum()
    return body[["key", "article", "stock"]]

def _fmt_from_stock_text(stock_txt: str) -> str | None:
    """Retourne '12x33' / '6x75' / '4x75' depuis la colonne Stock."""
    nb, vol = parse_stock(stock_txt)
    if pd.isna(nb) or pd.isna(vol): return None
    nb = int(nb); vol = float(vol)
    if nb == 12 and abs(vol - 0.33) <= VOL_TOL: return "12x33"
    if nb == 6  and abs(vol - 0.75) <= VOL_TOL: return "6x75"
    if nb == 4  and abs(vol - 0.75) <= VOL_TOL: return "4x75"
    return None

# ====================== Agrégation ventes -> prévisions ======================

def aggregate_forecast_by_format(
    df_sales: pd.DataFrame, window_days: float, horizon_j: int
) -> tuple[Dict[str, Dict[str, float]], Dict[str, Dict[str, Dict[str, float]]]]:
    """
    Retourne un double résultat:
      - fmt_totals[fmt] = {"bottles": ..., "cartons": ...}   (agrégé TOUS groupes)
      - by_group[group_key][fmt] = {"bottles": ..., "cartons": ...}
    Ici, group_key = **bucket** (famille | goût intrinsèque).
    """
    req = ["Stock", "Volume vendu (hl)", "GoutCanon"]
    if any(c not in df_sales.columns for c in req):
        return {}, {}

    tmp = df_sales.copy()
    tmp["fmt"] = tmp["Stock"].map(_fmt_from_stock_text)
    tmp = tmp.dropna(subset=["fmt"])
    parsed = tmp["Stock"].map(parse_stock)
    tmp[["nb_btl_cart", "vol_L"]] = pd.DataFrame(parsed.tolist(), index=tmp.index)

    tmp["vol_hL_per_btl"] = (tmp["vol_L"].astype(float) / 100.0)
    tmp["nb_btl_cart"] = pd.to_numeric(tmp["nb_btl_cart"], errors="coerce")
    tmp["v_hL_j"] = pd.to_numeric(tmp["Volume vendu (hl)"], errors="coerce") / max(float(window_days), 1.0)

    tmp["group"] = tmp["GoutCanon"].astype(str).str.strip()
    tmp = tmp.replace([np.inf, -np.inf], np.nan).dropna(
        subset=["vol_hL_per_btl", "nb_btl_cart", "v_hL_j"]
    )

    tmp["btl_j"] = np.where(tmp["vol_hL_per_btl"] > 0, tmp["v_hL_j"] / tmp["vol_hL_per_btl"], 0.0)
    tmp["carton_j"] = np.where(tmp["nb_btl_cart"] > 0, tmp["btl_j"] / tmp["nb_btl_cart"], 0.0)
    tmp["btl_h"] = horizon_j * tmp["btl_j"]
    tmp["carton_h"] = horizon_j * tmp["carton_j"]

    agg_fmt = tmp.groupby("fmt").agg(bottles=("btl_h", "sum"), cartons=("carton_h", "sum"))
    fmt_totals = {fmt: {"bottles": float(agg_fmt.loc[fmt, "bottles"]),
                        "cartons": float(agg_fmt.loc[fmt, "cartons"])} for fmt in agg_fmt.index}
    for k in ["12x33", "6x75", "4x75"]:
        fmt_totals.setdefault(k, {"bottles": 0.0, "cartons": 0.0})

    agg_ff = tmp.groupby(["group", "fmt"]).agg(bottles=("btl_h", "sum"), cartons=("carton_h", "sum"))
    by_group: Dict[str, Dict[str, Dict[str, float]]] = {}
    for (g, f), row in agg_ff.iterrows():
        by_group.setdefault(g, {})[f] = {"bottles": float(row["bottles"]),
                                         "cartons": float(row["cartons"])}
    for g in by_group:
        for f in ["12x33", "6x75", "4x75"]:
            by_group[g].setdefault(f, {"bottles": 0.0, "cartons": 0.0})

    return fmt_totals, by_group

# ====================== Famille + Goût intrinsèque + Formats ======================

def _pick_prod_column(df: pd.DataFrame) -> str:
    """Trouve la colonne qui contient le libellé produit (sans mapper)."""
    cand = ["produit","désignation","designation","libellé","libelle",
            "nom du produit","product","sku libellé","sku libelle","sku","item"]
    cols = {str(c).strip(): str(c).strip() for c in df.columns}
    norm = {re.sub(r"[^a-z0-9]+", " ", k.lower()).strip(): v for k,v in cols.items()}
    for k in cand:
        nk = re.sub(r"[^a-z0-9]+", " ", k.lower()).strip()
        if nk in norm:
            return norm[nk]
    return list(cols.values())[0]

def _family_from_produit(prod: str) -> str:
    p = _canon_txt(prod)
    if "inter" in p: return "inter"
    if "niko"  in p: return "niko"
    if "igeba" in p: return "igeba"
    return "fr"

# alias FR/EN pour détecter le goût intrinsèque
_FLAVOR_ALIASES = {
    "original": ["original","nature","classic"],
    "gingembre": ["gingembre","ginger"],
    "mangue passion": ["mangue passion","mango passion","mango-passion","mango  passion","mapa"],
    "menthe citron vert": ["menthe citron vert","menthe-citron vert","menthe-citron-vert","mint lime","mint-lime","mint & lime","mint and lime","mcv"],
    "pamplemousse": ["pamplemousse","grapefruit"],
    "infusion menthe poivrée": ["menthe poivree","menthe-poivree","peppermint"],
    "infusion mélisse": ["melisse","mélisse","lemonbalm","lemon balm","lemon-balm"],
    "infusion anis": ["anis","anise","star anise","anis etoile","anis étoilée"],
    "igeba pêche": ["igeba peche","igeba pêche","peach"],
}

def _extract_flavor(text: str) -> str:
    a = _canon_txt(text)
    best = None
    for canon, aliases in _FLAVOR_ALIASES.items():
        for al in aliases + [canon]:
            al_n = _canon_txt(al)
            if al_n and al_n in a:
                if best is None or len(al_n) > len(best[1]):
                    best = (canon, al_n)
    return best[0] if best else "(autre)"

def _bucket_key(family: str, flavor: str) -> str:
    return f"{family} | {flavor}"

def _article_applies_formats(article: str) -> Tuple[List[str], str]:
    """
    Formats cibles + unité par défaut.
    - '33' explicite -> ['12x33']
    - '75' explicite -> ['6x75']/'4x75' si précisé, sinon ['6x75','4x75']
    - Étiquettes INTER/NIKO/IGEBA sans précision -> 33 par défaut
    """
    a = _norm_txt(article)
    per = "carton" if any(w in a for w in ["carton", "caisse", "colis", "etui", "étui"]) else "bottle"

    if "12x33" in a or ("33" in a and "75" not in a): return ["12x33"], per
    if "6x75" in a: return ["6x75"], per
    if "4x75" in a: return ["4x75"], per
    if "75"  in a:  return ["6x75","4x75"], per
    if ("etiquette" in a or "étiquette" in a) and any(w in a for w in ["inter","niko","igeba"]):
        return ["12x33"], per
    return ["12x33","6x75","4x75"], per

def _targets_from_article(article: str, known_buckets: List[str]) -> List[str]:
    """
    Pour les étiquettes : renvoie le bucket 'famille | goût' ciblé.
    Pour articles génériques (capsules/cartons) -> [] (agrégat tous buckets).
    Si on ne détecte pas clairement le goût sur une étiquette, on retourne [] et on traitera comme 0.
    """
    a = _canon_txt(article)
    is_label = ("etiquette" in a or "étiquette" in a)

    if not is_label:
        return []  # générique

    # famille
    if   "inter" in a: fam = "inter"
    elif "niko"  in a: fam = "niko"
    elif "igeba" in a: fam = "igeba"
    else:               fam = "fr"

    flv = _extract_flavor(article)
    if flv == "(autre)":
        return []  # on préfère ne rien compter plutôt qu'agréger mal

    key = _bucket_key(fam, flv)
    return [key] if key in set(known_buckets) else []

# ====================== Sommes utilitaires ======================

def _sum_units_for_targets(
    targets: List[str], fmts: List[str], per: str,
    fmt_forecast: Dict[str, Dict[str, float]],
    ff_forecast: Dict[str, Dict[str, Dict[str, float]]]
) -> float:
    key = "bottles" if per == "bottle" else "cartons"
    total = 0.0
    if targets:  # spécifique bucket(s)
        for g in targets:
            for f in fmts:
                total += float(ff_forecast.get(g, {}).get(f, {}).get(key, 0.0))
    else:       # générique → agrégé tous buckets
        for f in fmts:
            total += float(fmt_forecast.get(f, {}).get(key, 0.0))
    return total

# ====================== Calcul de la table des besoins ======================

def compute_needs_table(
    df_conso: pd.DataFrame,
    df_stock: pd.DataFrame,
    *,
    forecast_fmt_H: Dict[str, Dict[str, float]],
    forecast_ff_H: Dict[str, Dict[str, Dict[str, float]]],
    forecast_fmt_ref: Dict[str, Dict[str, float]],
    forecast_ff_ref: Dict[str, Dict[str, Dict[str, float]]],
    force_labels: bool
) -> pd.DataFrame:
    """
    1) Détecte pour chaque article: formats + bucket(s) (famille|goût) visés
    2) Calcule un coef par unité depuis la période de conso (B2):
         coef = conso_total / unités_sur_période_B2
       (sauf articles "1 pour 1": coef = 1)
    3) Besoin(H) = coef × unités_prévues_sur_H
    """
    rows = []
    known_buckets = list(forecast_ff_H.keys())  # mêmes clés que ref

    for _, r in df_conso.iterrows():
        art = r["article"]; k = r["key"]
        conso_total = float(r["conso"])
        a_norm = _norm_txt(art)

        fmts, per = _article_applies_formats(art)
        targets = _targets_from_article(art, known_buckets)

        is_label    = ("etiquette" in a_norm or "étiquette" in a_norm)
        is_capsule  = ("capsule" in a_norm)
        is_transport_carton = ("carton" in a_norm and ("33" in a_norm or "75" in a_norm))

        # unités prévues
        if is_label and not targets:
            # étiquette non reconnue précisément → ne rien sur-agréger
            units_H = 0.0
            units_ref = 0.0
        else:
            units_H   = _sum_units_for_targets(targets, fmts, per, forecast_fmt_H,   forecast_ff_H)
            units_ref = _sum_units_for_targets(targets, fmts, per, forecast_fmt_ref, forecast_ff_ref)

        # Articles “1 pour 1”
        if force_labels and is_label:
            coef = 1.0
        elif is_capsule:
            coef = 1.0
        elif is_transport_carton:
            coef = 1.0
        else:
            coef = (conso_total / units_ref) if units_ref > 0 else 0.0

        besoin = coef * units_H

        rows.append({
            "key": k,
            "Article": art,
            "Unité": "par bouteille" if per == "bottle" else "par carton",
            "Besoin horizon": besoin
        })

    need_df = pd.DataFrame(rows)
    if need_df.empty:
        return pd.DataFrame(columns=["Article","Unité","Besoin horizon","Stock dispo","À acheter"])

    st_df = (df_stock[["key","stock"]].rename(columns={"stock":"Stock dispo"})
             if df_stock is not None else pd.DataFrame(columns=["key","Stock dispo"]))
    out = need_df.merge(st_df, on="key", how="left").fillna({"Stock dispo": 0.0})

    out["À acheter"] = np.maximum(out["Besoin horizon"] - out["Stock dispo"], 0.0)
    for c in ["Besoin horizon","Stock dispo","À acheter"]:
        out[c] = np.round(out[c], 0).astype(int)

    return out.drop(columns=["key"]).sort_values("À acheter", ascending=False).reset_index(drop=True)

# ====================== Calculs principaux ======================

# --- Construire le bucket FAMILLE|GOÛT intrinsèque (sans flavor_map) ---
prod_col = _pick_prod_column(df_raw)
df_ff = df_raw.copy()
df_ff["__prod"]  = df_ff[prod_col].astype(str)
df_ff["__fam"]   = df_ff["__prod"].map(_family_from_produit)
df_ff["__flav"]  = df_ff["__prod"].map(_extract_flavor)
df_ff["GoutCanon"] = df_ff.apply(lambda r: _bucket_key(r["__fam"], r["__flav"]), axis=1)

# Prévisions pour l’horizon courant (H)
forecast_fmt_H, forecast_ff_H = aggregate_forecast_by_format(
    df_ff, window_days=window_days, horizon_j=int(horizon_j)
)

# KPIs (étiquettes ≈ bouteilles)
b_33 = forecast_fmt_H.get("12x33", {}).get("bottles", 0.0)
b_75 = forecast_fmt_H.get("6x75", {}).get("bottles", 0.0) + forecast_fmt_H.get("4x75", {}).get("bottles", 0.0)
cartons_total = sum(v.get("cartons", 0.0) for v in forecast_fmt_H.values())

colA, colB, colC = st.columns([1.1, 1, 1])
with colA: kpi("Étiquettes à prévoir — 12x33", f"{b_33:.0f}")
with colB: kpi("Étiquettes à prévoir — 75cl",  f"{b_75:.0f}")
with colC: kpi("Cartons prévus (tous formats)", f"{cartons_total:.0f}")

# ====================== Lecture fichiers + résultat ======================

df_conso = None
conso_days = None
df_stockc = None
err_block = False

if conso_file is not None:
    try:
        df_conso, conso_days = read_consumption_xlsx(conso_file)
        st.success(f"Consommation: zone détectée ✅ — Période B2 = **{conso_days} jours**")
        with st.expander("Voir l’aperçu du fichier **Consommation**", expanded=False):
            st.dataframe(df_conso[["article", "conso", "per_hint"]], use_container_width=True, hide_index=True)
    except Exception as e:
        st.error(f"Erreur lecture consommation: {e}")
        err_block = True
else:
    st.info("Importer l’Excel **Consommation des articles** (bloc ci-dessus).")

if stock_file is not None:
    try:
        df_stockc = read_stock_xlsx(stock_file)
        st.success("Stocks: colonne 'Quantité virtuelle' détectée ✅")
        with st.expander("Voir l’aperçu du fichier **Stocks**", expanded=False):
            st.dataframe(df_stockc[["article", "stock"]], use_container_width=True, hide_index=True)
    except Exception as e:
        st.error(f"Erreur lecture stocks: {e}")
        err_block = True
else:
    st.info("Importer l’Excel **Stocks des articles** (bloc ci-dessus).")

st.markdown("---")

if (df_conso is not None) and (df_stockc is not None) and (not err_block):
    # Prévisions sur la même période que le fichier conso (référence B2)
    conso_days = int(conso_days or 30)
    forecast_fmt_ref, forecast_ff_ref = aggregate_forecast_by_format(
        df_ff, window_days=window_days, horizon_j=conso_days
    )

    result = compute_needs_table(
        df_conso, df_stockc,
        forecast_fmt_H=forecast_fmt_H, forecast_ff_H=forecast_ff_H,
        forecast_fmt_ref=forecast_fmt_ref, forecast_ff_ref=forecast_ff_ref,
        force_labels=force_labels
    )

    if result.empty:
        st.info("Aucun besoin calculé (vérifie les fichiers de consommation/stocks et les correspondances d’articles).")
        st.stop()

    total_buy = int(result["À acheter"].sum())
    nb_items  = int((result["À acheter"] > 0).sum())
    c1, c2 = st.columns(2)
    with c1: kpi("Articles à acheter (nb)", f"{nb_items}")
    with c2: kpi("Quantité totale à acheter (unités)", f"{total_buy:,}".replace(",", " "))

    st.subheader("Proposition d’achats (triée par 'À acheter' décroissant)")
    st.dataframe(
        result[["Article","Unité","Besoin horizon","Stock dispo","À acheter"]],
        use_container_width=True, hide_index=True,
        column_config={
            "Besoin horizon": st.column_config.NumberColumn(format="%d"),
            "Stock dispo":    st.column_config.NumberColumn(format="%d"),
            "À acheter":      st.column_config.NumberColumn(format="%d"),
        }
    )

    csv_bytes = result.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Exporter la proposition (CSV)",
        data=csv_bytes,
        file_name=f"achats_conditionnements_{int(horizon_j)}j.csv",
        mime="text/csv",
        use_container_width=True,
    )
else:
    st.info("Charge les deux fichiers pour obtenir la proposition d’achats.")
